import openpyxl as xl
import os
from openpyxl.chart import BarChart, Reference

file_path = "d:\Machine Learning\Python\intro-to-python\excelfile\\transactions.xlsx"
print(file_path)

if os.path.isfile(file_path):
    wb = xl.load_workbook(file_path)
    firstsheet = wb["Sheet1"]
    cell = firstsheet['a1']
    # row 1 col 1
    cell = firstsheet.cell(1, 1)
    print(cell.value)
    # Continue with your code
    for row in range(2, firstsheet.max_row + 1):
        # print(row)
        # 1 no row 3 no col 2 no row 3 no col
        cell = firstsheet.cell(row, 3)
        print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = firstsheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    for row in range(2, firstsheet.max_row + 1):
        cell = firstsheet.cell(row, 4)
        print(cell.value)
    values = Reference(firstsheet, min_col=2, min_row=1, max_row=5, max_col=2)
    chart = BarChart()
    chart.add_data(values)
    print(chart)
    res = firstsheet.add_chart(chart, "E2")
    print(res)
    wb.save("barchart.xlsx")
else:
    print("File not found: transactions.xlsx")
